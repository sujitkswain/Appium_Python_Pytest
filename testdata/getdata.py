
from pathlib import Path
import openpyxl


def get_data(filename, sheet_name):
    data_list = []
    wb = openpyxl.load_workbook(str(Path(__file__).parent) + "\\" + filename)
    sheet = wb[sheet_name]
    for i in range(2, sheet.max_row+1):
        data = {}
        for j in range(1, sheet.max_column+1):
            data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        data_list.append(data)
    return data_list
